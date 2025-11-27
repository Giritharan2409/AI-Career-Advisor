from flask import Flask, render_template, request, jsonify, send_file
import requests
import os
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO

app = Flask(__name__)

# Configure Gemini API
API_KEY = os.getenv('GEMINI_API_KEY', 'AIzaSyAfECe8OQyO6NAN6MK21hL2-evugNJ4v6Y')

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/chat', methods=['POST'])
def chat():
    try:
        data = request.json
        user_message = data.get('message', '')
        
        if not user_message:
            return jsonify({'error': 'No message provided'}), 400
        
        # Define valid keywords for AI and career-related topics
        valid_keywords = {
            'ai': ['artificial intelligence', 'machine learning', 'deep learning', 'neural networks', 'ai'],
            'data': ['data science', 'data analyst', 'data analysis', 'analytics', 'data preprocessing', 'dataset'],
            'career': ['career path', 'job', 'skill', 'certification', 'development', 'career', 'profession'],
            'programming': ['python', 'sql', 'programming', 'algorithms', 'models', 'code', 'coding'],
            'frameworks': ['tensorflow', 'pytorch', 'scikit-learn', 'pandas', 'numpy', 'framework'],
            'domains': ['nlp', 'computer vision', 'regression', 'classification', 'clustering', 'prediction'],
            'platform': ['exam', 'assessment', 'roadmap', 'learning', 'tutorial', 'resource', 'chatbot', 'advisor', 'guidance', 'book', 'course', 'certification']
        }
        
        # Check if the message contains any valid keywords
        message_lower = user_message.lower()
        is_valid = False
        
        for category, keywords in valid_keywords.items():
            for keyword in keywords:
                if keyword in message_lower:
                    is_valid = True
                    break
            if is_valid:
                break
        
        # If question is not valid, return error message
        if not is_valid:
            return jsonify({
                'response': 'This is not a valid platform to answer your question. I am specifically designed to help with AI and career-related topics such as machine learning, data science, career guidance, and learning resources. Please ask questions related to AI, data science, programming, or career development.'
            })
        
        # If valid, proceed with answering the question
        system_prompt = """You are an AI Career Advisor assistant. Help users with:
- AI and Machine Learning concepts
- Data Science and analytics
- Career guidance and learning paths
- Programming and technical skills
- Learning resources, tutorials, books, and certifications
- Information about this website's features (exam, assessment, roadmap)

Be encouraging and supportive. Answer in simple English with concise, focused responses."""
        
        summarized_prompt = f"""{system_prompt}

Answer this question in simple English. Keep your response short and directly related to the question asked. Do not add extra information.
        
Question: {user_message}

Provide a concise, simple answer:"""
        
        # Generate response using Gemini REST API
        url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash:generateContent?key={API_KEY}"
        
        answer_payload = {
            "contents": [
                { "role": "user", "parts": [ {"text": summarized_prompt} ] }
            ]
        }
        
        response = requests.post(url, json=answer_payload)
        response_data = response.json()
        
        # Debug: Print the response
        print("API Response:", response_data)
        
        # Extract the reply from the response
        if 'candidates' in response_data and len(response_data['candidates']) > 0:
            candidate = response_data['candidates'][0]
            if 'content' in candidate and 'parts' in candidate['content']:
                reply = candidate['content']['parts'][0].get('text', 'No text found')
            else:
                reply = 'No content in response'
        elif 'error' in response_data:
            reply = f"API Error: {response_data['error'].get('message', 'Unknown error')}"
        else:
            reply = 'Unexpected response format'
        
        return jsonify({
            'response': reply
        })
    
    except Exception as e:
        print(f"Error: {str(e)}")
        return jsonify({'error': f'API Error: {str(e)}'}), 500

@app.route('/exam')
def exam():
    return render_template('exam.html')

@app.route('/results')
def results():
    return render_template('results.html')

@app.route('/roadmap')
def roadmap():
    return render_template('roadmap.html')

@app.route('/generate_roadmap', methods=['POST'])
def generate_roadmap():
    try:
        data = request.json
        domain = data.get('domain', 'data_scientist')
        
        # Map domains to prompts
        domain_prompts = {
            'data_analyst': 'Generate a comprehensive learning roadmap for Data Analyst career path. Include: 1) Step-by-step learning path, 2) Free certification courses with links, 3) Recommended YouTube tutorial channels/playlists in English, 4) Essential books with links. Format the response as JSON with these keys: learningPath (array of steps), certifications (array with name and link), youtubeTutorials (array with title and link), books (array with title and link).',
            'data_scientist': 'Generate a comprehensive learning roadmap for Data Scientist career path. Include: 1) Step-by-step learning path, 2) Free certification courses with links, 3) Recommended YouTube tutorial channels/playlists in English, 4) Essential books with links. Format the response as JSON with these keys: learningPath (array of steps), certifications (array with name and link), youtubeTutorials (array with title and link), books (array with title and link).',
            'machine_learning': 'Generate a comprehensive learning roadmap for Machine Learning Engineer career path. Include: 1) Step-by-step learning path, 2) Free certification courses with links, 3) Recommended YouTube tutorial channels/playlists in English, 4) Essential books with links. Format the response as JSON with these keys: learningPath (array of steps), certifications (array with name and link), youtubeTutorials (array with title and link), books (array with title and link).'
        }
        
        prompt = domain_prompts.get(domain, domain_prompts['data_scientist'])
        
        # Generate response using Gemini REST API
        url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash:generateContent?key={API_KEY}"
        
        payload = {
            "contents": [
                { "role": "user", "parts": [ {"text": prompt} ] }
            ],
            "generationConfig": {
                "response_mime_type": "application/json"
            }
        }
        
        response = requests.post(url, json=payload)
        response_data = response.json()
        
        # Extract the reply from the response
        if 'candidates' in response_data and len(response_data['candidates']) > 0:
            candidate = response_data['candidates'][0]
            if 'content' in candidate and 'parts' in candidate['content']:
                reply = candidate['content']['parts'][0].get('text', 'No text found')
                # Parse the JSON response
                try:
                    roadmap_data = json.loads(reply)
                    return jsonify({
                        'roadmap': roadmap_data
                    })
                except json.JSONDecodeError:
                    return jsonify({'error': f'Failed to parse roadmap: {reply}'}), 500
            else:
                return jsonify({'error': 'No content in response'}), 500
        elif 'error' in response_data:
            return jsonify({'error': f"API Error: {response_data['error'].get('message', 'Unknown error')}"}), 500
        else:
            return jsonify({'error': 'Unexpected response format'}), 500
    
    except Exception as e:
        print(f"Error: {str(e)}")
        return jsonify({'error': f'API Error: {str(e)}'}), 500

@app.route('/download_roadmap', methods=['POST'])
def download_roadmap():
    """Generate and download roadmap as Excel file"""
    try:
        data = request.json
        domain = data.get('domain', 'data_scientist')
        roadmap_data = data.get('roadmap', {})
        
        # Domain names mapping
        domain_names = {
            'data_analyst': 'Data Analyst',
            'data_scientist': 'Data Scientist',
            'machine_learning': 'Machine Learning Engineer'
        }
        domain_name = domain_names.get(domain, 'Data Scientist')
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = domain_name
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="0072FF", end_color="0072FF", fill_type="solid")
        section_font = Font(bold=True, color="FFFFFF", size=11)
        section_fill = PatternFill(start_color="00C6FF", end_color="00C6FF", fill_type="solid")
        alignment = Alignment(wrap_text=True, vertical="top")
        
        # Add title
        ws['A1'] = f'{domain_name} - Learning Roadmap'
        ws['A1'].font = Font(bold=True, size=14, color="0072FF")
        ws.merge_cells('A1:C1')
        ws['A1'].alignment = Alignment(horizontal="center")
        
        current_row = 3
        
        # Add Learning Path section
        ws[f'A{current_row}'] = 'Learning Path'
        ws[f'A{current_row}'].font = section_font
        ws[f'A{current_row}'].fill = section_fill
        ws.merge_cells(f'A{current_row}:C{current_row}')
        current_row += 1
        
        if roadmap_data.get('learningPath'):
            for idx, step in enumerate(roadmap_data['learningPath'], 1):
                ws[f'A{current_row}'] = f'Step {idx}'
                ws[f'B{current_row}'] = step
                ws[f'A{current_row}'].alignment = alignment
                ws[f'B{current_row}'].alignment = alignment
                ws.row_dimensions[current_row].height = None
                current_row += 1
        
        current_row += 1
        
        # Add Free Certifications section
        ws[f'A{current_row}'] = 'Free Certifications'
        ws[f'A{current_row}'].font = section_font
        ws[f'A{current_row}'].fill = section_fill
        ws.merge_cells(f'A{current_row}:C{current_row}')
        current_row += 1
        
        # Add headers for certifications
        ws[f'A{current_row}'] = 'Certification Name'
        ws[f'B{current_row}'] = 'Link'
        for cell in [ws[f'A{current_row}'], ws[f'B{current_row}']]:
            cell.font = header_font
            cell.fill = header_fill
        current_row += 1
        
        if roadmap_data.get('certifications'):
            for cert in roadmap_data['certifications']:
                ws[f'A{current_row}'] = cert.get('name', '')
                ws[f'B{current_row}'] = cert.get('link', '')
                ws[f'A{current_row}'].alignment = alignment
                ws[f'B{current_row}'].alignment = alignment
                current_row += 1
        
        current_row += 1
        
        # Add YouTube Tutorials section
        ws[f'A{current_row}'] = 'YouTube Tutorials'
        ws[f'A{current_row}'].font = section_font
        ws[f'A{current_row}'].fill = section_fill
        ws.merge_cells(f'A{current_row}:C{current_row}')
        current_row += 1
        
        # Add headers for tutorials
        ws[f'A{current_row}'] = 'Tutorial Title'
        ws[f'B{current_row}'] = 'Link'
        for cell in [ws[f'A{current_row}'], ws[f'B{current_row}']]:
            cell.font = header_font
            cell.fill = header_fill
        current_row += 1
        
        if roadmap_data.get('youtubeTutorials'):
            for tutorial in roadmap_data['youtubeTutorials']:
                ws[f'A{current_row}'] = tutorial.get('title', '')
                ws[f'B{current_row}'] = tutorial.get('link', '')
                ws[f'A{current_row}'].alignment = alignment
                ws[f'B{current_row}'].alignment = alignment
                current_row += 1
        
        current_row += 1
        
        # Add Books section
        ws[f'A{current_row}'] = 'Recommended Books'
        ws[f'A{current_row}'].font = section_font
        ws[f'A{current_row}'].fill = section_fill
        ws.merge_cells(f'A{current_row}:C{current_row}')
        current_row += 1
        
        # Add headers for books
        ws[f'A{current_row}'] = 'Book Title'
        ws[f'B{current_row}'] = 'Link'
        for cell in [ws[f'A{current_row}'], ws[f'B{current_row}']]:
            cell.font = header_font
            cell.fill = header_fill
        current_row += 1
        
        if roadmap_data.get('books'):
            for book in roadmap_data['books']:
                ws[f'A{current_row}'] = book.get('title', '')
                ws[f'B{current_row}'] = book.get('link', '')
                ws[f'A{current_row}'].alignment = alignment
                ws[f'B{current_row}'].alignment = alignment
                current_row += 1
        
        # Set column widths
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 20
        
        # Save to BytesIO object
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Return file for download
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'{domain_name}_Learning_Roadmap.xlsx'
        )
    
    except Exception as e:
        print(f"Error: {str(e)}")
        return jsonify({'error': f'Failed to generate Excel: {str(e)}'}), 500

@app.route('/test-api')
def test_api():
    """Test endpoint to verify Gemini API"""
    try:
        # First, let's list available models
        list_url = f"https://generativelanguage.googleapis.com/v1beta/models?key={API_KEY}"
        list_response = requests.get(list_url)
        models_data = list_response.json()
        
        # Then test with gemini-2.5-flash if available
        test_url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash:generateContent?key={API_KEY}"
        payload = {
            "contents": [
                { "role": "user", "parts": [ {"text": "Say hello"} ]
                }
            ]
        }
        test_response = requests.post(test_url, json=payload)
        
        return jsonify({
            "available_models": models_data,
            "test_response": test_response.json()
        })
    except Exception as e:
        return jsonify({"error": str(e)})

@app.route('/list-models')
def list_models():
    """List all available models"""
    try:
        url = f"https://generativelanguage.googleapis.com/v1beta/models?key={API_KEY}"
        response = requests.get(url)
        return jsonify(response.json())
    except Exception as e:
        return jsonify({"error": str(e)})

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)